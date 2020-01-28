
# -*- coding: utf-8 -*-

import openpyxl
import csv
import re
import glob
import os


files = glob.glob(os.getcwd() + "/source/*.tsv")
for fl in files:
    source_content = []
    bookname = fl.split("/")[-1].split(".")[0]
    filename = fl.split("/")[-1]
    s_path = glob.glob(os.getcwd() + "/source/" + str(filename))
    t_path = glob.glob(os.getcwd() + "/target/" + str(bookname) + ".xlsx")

    # ----------- Target File ----------------#
    wb_obj = openpyxl.load_workbook(t_path[0])
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    row_count = 1
    # ---------- Source File -----------------#
    label_file = (bookname + ".tsv")
    # print (label_file)
    with open(label_file, 'w', encoding = 'utf-8') as tsv_file:
        twriter = csv.writer(tsv_file, delimiter='\t')
        with open(s_path[0], 'r', encoding = 'utf-8') as tsvfile:
            reader = csv.reader(tsvfile, delimiter='\t')
            for rows in reader:
                dic = {}
                eng_ocr_note = rows[8]
                hin_ocr_note = sheet_obj.cell(row=row_count, column=4).value
                if hin_ocr_note == '':
                    pass
                else:
                    try:
                        rep_br = hin_ocr_note.replace('$', '<br>')
                    except:
                        print(hin_ocr_note)
                    rep_see = re.sub(r'(\(देखें:.*?\))', "देखें:", rep_br)
                find_see = re.findall(r'(\(See:.*?\]\]\))', rows[8])
                if find_see:
                    for see in find_see:
                        if see == '':
                            pass
                        else:
                            dic[see] = "देखें:"
                    for k, v in dic.items():
                        rep_see = rep_see.replace(v, k, 1)
                    rep_see1 = re.sub(r'See', 'देखें', rep_see)
                    rep_and = re.sub(r'and', 'और', rep_see1)
                    # print(rep_and)
                    twriter.writerow([rows[0], rows[1], rows[2], rows[3], rows[4], rows[5], rows[6], rows[7], rep_and])
                    row_count += 1
                else:
                    # print(rep_see)
                    twriter.writerow([rows[0], rows[1], rows[2], rows[3], rows[4], rows[5], rows[6], rows[7], rep_see])
                    row_count += 1
            print("Completed")
